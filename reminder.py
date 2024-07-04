import os
from tkinter import *
import pandas as pd
from tkinter import filedialog, messagebox
import schedule
import time
from datetime import datetime
from twilio.rest import Client
import smtplib
import threading
import openpyxl
import json
# import keys
from playsound import playsound
import json

# global latest_file_number
# latest_file_number = 7


# import os

# Define the path to the JSON file
data_file_path = 'data.json'

# Function to save data to a JSON file
def save_data(data, file_path):
    try:
        with open(file_path, 'w') as file:
            json.dump(data, file)
        print(f"Data saved successfully to '{file_path}'")
    except Exception as e:
        print(f"An error occurred while saving data: {e}")

# Function to load data from a JSON file
def load_data(file_path):
    if os.path.exists(file_path):
        try:
            with open(file_path, 'r') as file:
                data = json.load(file)
            print(f"Data loaded successfully from '{file_path}'")
            return data
        except Exception as e:
            print(f"An error occurred while loading data: {e}")
            return None
    else:
        print(f"No data file found at '{file_path}'. Starting with empty data.")
        return None


# Example usage: modify the data

# print(f"Current counter value: {data['counter']}")

# Save the data before the program exits




def sendMail(date, name):  
    print("msg => "+ str(date) + str(name))
    server = smtplib.SMTP('smtp.gmail.com',587)
    status_code, response=server.ehlo()
    print(f"ehlo >>>>> {status_code} and {response}")
    status_codetls, responsetls=server.starttls()
    print(f"starttls >>>>> {status_codetls} and {responsetls}")

    subject="Missing Payment Alert"
    message = "Subject : {}\n\n{}".format(subject, "these user of the payment are not completed => \nnames are => "+ str(name) +'\ndates are => '+ str(date) )

    try:
        status_codel, responsel=server.login('guptagopal18082003@gmail.com', 'bwmczqycwmydrddq')
        print(f"these is login >>>>>  {status_codel} and {responsel}")
        print("Login successful")
        server.sendmail('guptagopal18082003@gmail.com', ['vg870779@gmail.com', 'vg660856@gmail.com'], message)
        print("mail sended")
    except smtplib.SMTPAuthenticationError as e:
        print(f"Login failed: {e}")
    finally:
        server.quit()

def playAlarm():
    time.sleep(1)
    playsound('C:\\Users\\Gopal gupta\\OneDrive\\Desktop\\tescom project 2\\ringtone.mp3')
    messagebox.showwarning("Warning", "please cheack unpaid payment.")

# def sendSMS(date, name):
#     # print("sms sending on... ",message.sid)
#     client = Client(keys.account_sid, keys.auth_token)
#     message = client.messages.create(
#       body="these user of the payment are not completed => "+ str(name) + str(date),   
#       from_=keys.twilio_number,
#       to=keys.my_phone_number
#     )
#     print("sms sended on ",message.sid)

def run_schedule():
    print("schedule 3")
    while True:
        schedule.run_pending()
        time.sleep(1)

def delete_file():
    try:
        # Check if the file exists
        if os.path.exists(file_path):
            # Remove read-only attribute if set
            os.chmod(file_path, 0o777)
            
            # Delete the file
            os.remove(file_path)
            print(f"The file '{file_path}' has been deleted successfully.")
        else:
            print(f"The file '{file_path}' does not exist.")
    except PermissionError:
        print(f"PermissionError: Unable to delete '{file_path}'. The file might be open in another application or you might not have the necessary permissions.")
    except Exception as e:
        print(f"An error occurred while deleting the file: {e}")

def updateXlData():
    global file_path
    user_input = entry.get()
    workbook = openpyxl.load_workbook(file_path)
    print(file_path.index)
    sheet = workbook['Sheet1']
    print("sheet.max_column => ",sheet.max_column)
    print("user_input => ", user_input)
    # Define the column indexes
    name_col = 1  # Column A for user names (1-based index)
    payment_status_col = sheet.max_column  # Column D for payment status (1-based index)

    # Iterate through the rows and update cells where payment status is "Not Done" and name is "Person8"
    for row in range(2, sheet.max_row + 1):  # Assuming the first row is the header
        name = sheet.cell(row=row, column=name_col).value
        payment_status = sheet.cell(row=row, column=payment_status_col).value
        if name == user_input and payment_status == "not done":
            # Update the cell to "Done"
            sheet.cell(row=row, column=payment_status_col).value = "done"

    # Save the workbook
    # C:/Users/Gopal gupta/OneDrive/Desktop/date.xlsx
    # file_path.rsplit("/")[-1] => date.xlsx
    # latest_file_number=latest_file_number+1
    # df.to_excel(f"date{latest_file_number+1}.xlsx", index=False)
    # data['counter'] += 1
    # print("load_data(\"data.json\") => ",load_data("data.json")-1)
    data = load_data(data_file_path)
    print("data => ", data)
    save_data(data+1, "data.json")
    workbook.save(f"date{data}.xlsx")
    print("done updates")
    delete_file()
    file_path=f"date{data}.xlsx"
    status_label2.config(text=f"latest File Name is => date{load_data("data.json")-1}")
    
    
    # return 0 
    # place always name column first in the excel sheet and as well as Paymemt column should be place in
    
    # # Load the workbook and select the sheet
    # workbook = openpyxl.load_workbook('example.xlsx')
    # sheet = workbook['Sheet1']
    
    # # Define the column indexes
    # name_col = 1  # Column A for user names (1-based index)
    # payment_status_col = 4  # Column D for payment status (1-based index)
    
    # # Iterate through the rows and update cells where payment status is "Not Done" and name is "Person8"
    # for row in range(2, sheet.max_row + 1):  # Assuming the first row is the header
    #     name = sheet.cell(row=row, column=name_col).value
    #     payment_status = sheet.cell(row=row, column=payment_status_col).value
    #     if name == "Person8" and payment_status == "Not Done":
    #         # Update the cell to "Done"
    #         sheet.cell(row=row, column=payment_status_col).value = "Done"
    
    # # Save the workbook
    # workbook.save('example_updated.xlsx')

    # wb = openpyxl.load_workbook(file_path)
    # ws = wb['Sheet1']


    # Get the value from the entry widget
    # print("file_path 1 => ", file_path)
    # # Check if the name exists in the 'Name' column
    # # if user_input in df['Name'].values:
    # if "person8" in df['Name'].values:
    #     print("file_path 2 => ", file_path)
    #     # Update the 'Payment' column to 'Done' for users with the entered name
    #     df.loc[df['Name'] == "person8", 'Payment'] = 'done'
    #     # Save the updated DataFrame back to the Excel file
    #     df.to_excel(file_path, index=False)
    #     print("file_path 3 => ", file_path)

    #     # Show a success message
    #     messagebox.showinfo("Success", f"Payment updated to 'Done' for {user_input}.")
    # else:
    #     # Show an error message if the name is not found
    #     messagebox.showerror("Error", f"Name {user_input} not found.")    
    # # Print the value to the console
    # messagebox.showinfo("Message", user_input)
    # print("user_input => ", user_input)

def start_scheduler():
    print("schedule 4")
    # Run the schedule in a separate thread to avoid blocking the Tkinter mainloop
    scheduler_thread = threading.Thread(target=run_schedule)
    scheduler_thread.daemon = True  # Ensures the thread exits when the main program does
    scheduler_thread.start()
    print("schedule 5")

def select_file():
    global file_path
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if file_path:
        excel_path.set(file_path)
        status_label.config(text=f"Selected file: {file_path}")
        status_label2.config(text=f"latest File Name is => date{load_data("data.json")-1}")

def dateToMicroseconds(year=0, month=0, day=0, hour=0, minute=0, second=0):
    dt = datetime(year, month, day, hour, minute, second)
    milliseconds = int(round(dt.timestamp() * 1000))
    # print(milliseconds)
    return milliseconds


def readXLdata():    
    file_path = excel_path.get()
    # print("file_path.rsplit(\"/\") => ",file_path.rsplit("/")[-1])
    if file_path:
        try:
            # Read the Excel file
            global df
            df = pd.read_excel(file_path, dtype={"Date": str})
            date = df["Date"].values
            print(date)
            print("____________________________________________________________________-")
            # Filter rows where the Payment status is 'not done' and select the 'Name' column
                #  and as well as
            # those who have payment atleast before 7 days
            not_done_payments_with_date = df[df['Payment'] == 'not done']['Date']
            not_done_payments_with_name = df[df['Payment'] == 'not done']['Name']
            # df['Date'] = pd.to_datetime(df['Date'])
            # datess=[]
            # not_done_payments = df[(df['Payment'] == 'not done') & (df['Date'] > datetime.now())]
            
            # # Iterate over the filtered DataFrame rows
            # for index, row in not_done_payments.iterrows():
            #     datess.append(row['Date'])
            
            # print(datess)

            # print("datess 1 => ", datess)


            # if (pd.to_datetime(not_done_payments_with_date)>datetime.now()):
            #     datess.append(not_done_payments_with_date)  
            
            # datess.append()
            # not_done_payments_with_name = df[df['Payment'] == 'not done']['Name']
            # Convert the Date column to datetime
            # df['Date'] = pd.to_datetime(df['Date'], format='%d-%m-%Y-%H-%M-%S')
            
            # Filter the DataFrame to include only rows with Date greater than the current date
            # name_of_user_not_done_payment= df[df['Date'] > datetime.now()]['Name']
            # print("name_of_user_not_done_payment => ", name_of_user_not_done_payment)
            # filtered_dates.strftime('%d-%m-%Y-%H-%M-%S')
            # print(day,"-",month,"-",year)
            # print("filtered_dates => ", filtered_dates.strftime('%d-%m-%Y-%H-%M-%S'))
            # print("filtered_dates username => ", df[df['Date']==filtered_dates]['Name'])
            # print("type of date column =>>>> ", type(int(df["Date"])))
            # print("not_done_payments_with_date => ", not_done_payments_with_date)
            # Print the names of users who have not done the payment
            dates = []
            names = []
            for xldate in not_done_payments_with_date:
                date, month, year, hour, minute, second=map(int, xldate.split("-")) # split day time
                print("date, month, year, hour, minute, second => ",date,"d", month,"m", year,"y", hour,"h", minute,"mint", second,"sec")
                dueDate = dateToMicroseconds(year, month, date, hour, minute, second)
                # currnetTime
                c = datetime.now()
                current_time = c.strftime('%d-%m-%Y-%H-%M-%S')
                Current_date, Current_month, Current_year, Current_hour, Current_minute, Current_second=map(int, current_time.split("-"))
                print("current ||| date, month, year, hour, minute, second => ", Current_date, Current_month, Current_year, Current_hour, Current_minute, Current_second)
                currentTime = dateToMicroseconds(Current_year, Current_month, Current_date, Current_hour, Current_minute, Current_second)
                # print(type(dueDate))
                # if the payment date before the current date and 7days before the current date
                # 7day => into millisecond => 604800000
                if dueDate>=currentTime and currentTime+604800000>dueDate:   
                    # print("xldate => \n", xldate)
                    df['Date'] = df['Date'].astype(str)
                    dates.append(xldate)
                    print("dates => ",dates)    # ['03-07-2024-00-00-00'] => 
                    # print("print username with it's date => ", df[xldate]['Name'])
                    # print("datess => ",datess)
                    print("____________________start from________________________")
                    for i in dates:
                        matched_rows = df[df['Date'] == i]

                    if not matched_rows.empty:
                        for name in matched_rows['Name']:
                            print("names are => ", name)
                            names.append(name)
                    else:
                        print("No user found with the specified date.")
                    print("these date are payment are not done as well as under next 7 day pending payment =>>>>>> ",date)
                    # send mail/sms/acknowledgment
                    
                    print("acknowledgment senting...")

                else : 
                    print("payment due date expired => ", dueDate, "payment day not within the range of 7 day day from currnet day")

            # print("not_done_payments_with_date.split("-")",not_done_payments_with_date.split("-"))

        except Exception as e:
            messagebox.showerror("Error", str(e))
    else:
        messagebox.showwarning("Warning", "Please select an Excel file first.")
    print("start mail of pending payment user")
    start_scheduler()
    # morning alert
    # not_done_payments_with_username
    print("not_done_payments_with_name =>>", not_done_payments_with_name)
    print("file_path 1 => ", file_path)

    # give the alert 3 time in a day, then you never forget and don't miss any of the payments
    schedule.every().day.at("11:00").do(sendMail, date=dates, name=names)
    schedule.every().day.at("11:00").do(playAlarm)

    schedule.every().day.at("13:00").do(sendMail, date=dates, name=names)
    schedule.every().day.at("13:00").do(playAlarm)
    # playsound('ringtone.mp3')
    
    # evening alert
    schedule.every().day.at("18:00").do(sendMail, date=dates, name=names)
    schedule.every().day.at("18:00").do(playAlarm)
    # playsound('ringtone.mp3')


    # send mail to pay the payment

# def submit_date():
#     date_input = date_entry.get()
#     # Validate the date format
#     try:
#         day, month, year = map(int, date_input.split('-'))
#         if len(date_input) == 10 and date_input[2] == '-' and date_input[5] == '-':
#             messagebox.showinfo("Success", f"Date entered: {date_input}")
#         else:
#             raise ValueError
#     except ValueError:
#         messagebox.showerror("Error", "Invalid date format. Please enter date as DD-MM-YYYY.")

# Create the main window
root = Tk()
root.title("Date Input")
root.geometry("1000x900")
root.title("Payment Reminder")
root.state("zoomed")
# Create and place the label

excel_path = StringVar()

frame = Frame(root)
frame.pack(expand=True)

select_button = Button(frame, text="Select Excel File", command=select_file)
select_button.pack(pady=10, ipadx=20, ipady=10)

# Create and place the entry widget

# Create and place the submit button
submit_button = Button(frame, text="Submit", command=readXLdata)  #if the data is to be updateded in xls sheet again call readXLdata
submit_button.pack(pady=10)


#Initialize a Label to display the User Input
label=Label(frame, text="", font=("Courier 22 bold"))
label.pack()

#Create an Entry widget to accept User Input
entry= Entry(frame, width= 40)
# entry.focus_set()
entry.pack()

submit_button = Button(frame, text="update on excel", command=updateXlData)
submit_button.pack(pady=10)

status_label = Label(frame, text="No file selected", anchor="center")
status_label.pack(pady=20)

status_label2 = Label(frame, text=f"latest File Name is => date{load_data("data.json")-1}", anchor="center")
status_label2.pack(pady=20)

# Start the Tkinter event loop
root.mainloop()


# date should be in 28-06-2024-00-00-00 this format so used in excel this code =TEXT(A1, "dd-mm-yyyy") & "-00-00-00"
